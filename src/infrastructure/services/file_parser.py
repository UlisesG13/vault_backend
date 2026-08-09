"""Parseo de archivos CSV y Excel (.xlsx) a filas de diccionarios."""
import csv
import io

from ...domain.exceptions import ValidationError
from ...domain.interfaces.services import IFileParser


def _norm(header: str | None) -> str:
    """Normaliza un encabezado: minúsculas, sin espacios extremos, espacios -> guion bajo."""
    return (header or "").strip().lower().replace(" ", "_")


class FileParserService(IFileParser):
    def parse(self, content: bytes, filename: str) -> list[dict[str, str]]:
        name = filename.lower()
        if name.endswith(".csv"):
            return self._parse_csv(content)
        if name.endswith(".xlsx"):
            return self._parse_xlsx(content)
        raise ValidationError("Formato no soportado: usa un archivo .csv o .xlsx")

    @staticmethod
    def _parse_csv(content: bytes) -> list[dict[str, str]]:
        try:
            text = content.decode("utf-8-sig")  # tolera BOM de Excel
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        reader = csv.DictReader(io.StringIO(text))
        if reader.fieldnames is None:
            raise ValidationError("El CSV no tiene encabezados")
        rows: list[dict[str, str]] = []
        for raw in reader:
            rows.append({_norm(k): (v or "") for k, v in raw.items() if k is not None})
        return rows

    @staticmethod
    def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise ValidationError(
                "Soporte de Excel no disponible: falta la dependencia 'openpyxl'"
            ) from exc

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValidationError("El Excel está vacío")

        headers = [_norm(str(h)) if h is not None else "" for h in header_row]
        rows: list[dict[str, str]] = []
        for values in rows_iter:
            if values is None or all(v is None for v in values):
                continue  # ignora filas totalmente vacías
            row = {
                headers[i]: ("" if values[i] is None else str(values[i]))
                for i in range(len(headers))
                if headers[i]
            }
            rows.append(row)
        workbook.close()
        return rows
